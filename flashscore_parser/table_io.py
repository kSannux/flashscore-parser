from __future__ import annotations

import json
import re
from copy import copy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText, InlineFont, TextBlock
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries

from flashscore_parser.table_io_xlwings import fill_template_with_excel

PLACEHOLDER_RE = re.compile(r"\{((?:[^{}]|\{[^{}]*\})+)\}")
CELL_REF = r"\$?[A-Z]{1,3}\$?\d+"
FORMAT_ATTR_RE = re.compile(
    rf"^\s*.+?\s*\?\s*{CELL_REF}\s*:\s*"
    rf"(?:.+?\s*\?\s*{CELL_REF}\s*:\s*)*{CELL_REF}\s*$"
)
FORMAT_RULE_RE = re.compile(
    rf"^\s*(?P<condition>.+?)\s*\?\s*"
    rf"(?P<cell>{CELL_REF})\s*:\s*(?P<remaining>.+)$"
)
CELL_REF_RE = re.compile(rf"^{CELL_REF}$")
MISSING_PLACEHOLDER_VALUE = 0
NUMERIC_ODDS_NAMES = {
    "over",
    "under",
    "asian-1",
    "asian-2",
    "favorite-asian",
    "outsider-asian",
    "european-1",
    "european-x",
    "european-2",
}

@dataclass(frozen=True)
class XlsxTemplate:
    headers: list[str]
    row_templates: list[list[Any]]
    sheet_name: str
    header_row: int
    template_start_row: int

@dataclass(frozen=True)
class RenderedArrayFormula:
    formula: str
    array_ref: str

    def shifted_ref(self, row_offset: int) -> str:
        def replace(match: re.Match[str]) -> str:
            column = match.group("column")
            row = match.group("row")
            if row.startswith("$"):
                return f"{column}{row}"
            return f"{column}{int(row) + row_offset}"

        return re.sub(
            r"(?P<column>\$?[A-Z]{1,3})(?P<row>\$?\d+)",
            replace,
            self.array_ref,
        )

    def cells(self, row_offset: int) -> set[tuple[int, int]]:
        min_column, min_row, max_column, max_row = range_boundaries(self.shifted_ref(row_offset))
        return {
            (row, column)
            for row in range(min_row, max_row + 1)
            for column in range(min_column, max_column + 1)
        }


WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}


def ensure_workbook_path(path: str | Path) -> Path:
    workbook_path = Path(path)
    if workbook_path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise RuntimeError(f"Поддерживаются только .xlsx и .xlsm файлы: {workbook_path}")
    return workbook_path


def keep_vba(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def close_workbook(workbook: Any) -> None:
    """Закрывает также ZIP-архив VBA, который openpyxl оставляет открытым."""
    try:
        workbook.close()
    finally:
        vba_archive = getattr(workbook, "vba_archive", None)
        if vba_archive is not None:
            vba_archive.close()


def cell_text(value: Any) -> str:
    if isinstance(value, ArrayFormula):
        return value.text
    return "" if value is None else str(value)

def cell_has_placeholder(value: str) -> bool:
    return PLACEHOLDER_RE.search(cell_text(value)) is not None

def row_has_value(row: tuple[Cell, ...]) -> bool:
    return any(cell_has_placeholder(cell.value) for cell in row)

def is_special_name(name: str) -> bool:
    if name in {"repeat"}:
        return True
    return False
        
def is_format_attr(attr: str) -> re.Match[str] | None:
    return FORMAT_ATTR_RE.fullmatch(attr)


def split_attrs(attrs: str) -> list[str]:
    parts: list[str] = []
    start = 0
    nesting = 0

    for index, character in enumerate(attrs):
        if character == "{":
            nesting += 1
        elif character == "}":
            nesting = max(0, nesting - 1)
        elif character == ";" and nesting == 0:
            parts.append(attrs[start:index].strip())
            start = index + 1

    parts.append(attrs[start:].strip())
    return parts


def parse_format_attr(attr: str) -> tuple[list[tuple[str, str]], str]:
    remaining = attr.strip()
    rules: list[tuple[str, str]] = []

    while True:
        match = FORMAT_RULE_RE.fullmatch(remaining)
        if match is None:
            break
        rules.append((match.group("condition"), match.group("cell")))
        remaining = match.group("remaining").strip()

    if not rules or CELL_REF_RE.fullmatch(remaining) is None:
        raise RuntimeError(f"Некорректный атрибут форматирования: {attr}")
    return rules, remaining


def split_format_attr(placeholder: str) -> tuple[str, str | None]:
    key = placeholder.strip()
    if ":" not in key:
        return key, None

    name, attrs = key.split(":", 1)
    regular_attrs: list[str] = []
    format_attr: str | None = None

    for attr in split_attrs(attrs):
        if is_format_attr(attr):
            if format_attr is not None:
                raise RuntimeError(f"В плейсхолдере указан второй атрибут форматирования: {{{key}}}")
            format_attr = attr
        else:
            regular_attrs.append(attr)

    if format_attr is None:
        return key, None
    if not regular_attrs:
        return name, format_attr
    return f"{name}:{';'.join(regular_attrs)}", format_attr


def font_to_inline(font: Any) -> InlineFont:
    return InlineFont(
        rFont=font.name,
        charset=font.charset,
        family=font.family,
        b=font.b,
        i=font.i,
        strike=font.strike,
        outline=font.outline,
        shadow=font.shadow,
        condense=font.condense,
        extend=font.extend,
        color=copy(font.color),
        sz=font.sz,
        u=font.u,
        vertAlign=font.vertAlign,
        scheme=font.scheme,
    )


def resolve_format_cell(attr: str, values: dict[str, Any], worksheet: Worksheet) -> Cell:
    def replace_condition_placeholder(placeholder_match: re.Match[str]) -> str:
        value = resolve_placeholder(values, placeholder_match.group(1))
        return repr(value)

    rules, default_cell = parse_format_attr(attr)
    for raw_condition, cell_ref in rules:
        condition = PLACEHOLDER_RE.sub(replace_condition_placeholder, raw_condition)
        try:
            condition_result = bool(eval(condition, {"__builtins__": {}}, values))
        except (NameError, SyntaxError, TypeError, ValueError) as error:
            raise RuntimeError(f"Не удалось вычислить условие форматирования '{condition}': {error}") from error

        if condition_result:
            return worksheet[cell_ref.replace("$", "")]

    return worksheet[default_cell.replace("$", "")]


def resolve_format_attr(attr: str, values: dict[str, Any], worksheet: Worksheet) -> InlineFont:
    return font_to_inline(resolve_format_cell(attr, values, worksheet).font)
        
def is_h2h_name(name: str):
    if name in {"home", "away", "h2h"}:
        return True
    return False

def is_valid_odds_parameter(name: str, parameter: str) -> bool:
    if name in NUMERIC_ODDS_NAMES:
        try:
            value = Decimal(parameter)
        except InvalidOperation:
            return False
        return value.is_finite() and value % Decimal("0.25") == 0
    if name == "correct":
        return re.fullmatch(r"\d+:\d+", parameter) is not None
    if name == "ht-ft":
        return re.fullmatch(r"[12X]/[12X]", parameter) is not None
    return False

def is_valid_missing_odds_placeholder(name: str, attrs: str) -> bool:
    if "-" not in attrs:
        return False
    parameter, state = attrs.rsplit("-", 1)
    return state in {"prev", "final"} and is_valid_odds_parameter(name, parameter)

def resolve_placeholder(values: dict[str, Any], placeholder: str) -> Any:
    key = placeholder.strip()
    if key in values:
        value = values[key]
        if isinstance(value, dict) and not value:
            return MISSING_PLACEHOLDER_VALUE
        if isinstance(value, (list, tuple)) and tuple(value) == (0.0, 0.0):
            return MISSING_PLACEHOLDER_VALUE
        return value

    if ":" not in key:
        return ""

    name, attrs = key.split(":", 1)
    if is_special_name(name):
        odd = resolve_placeholder(values, attrs)
        if odd == "":
            return ""
        repeat_values = values["repeat"].setdefault(key, {})
        repeat_values[odd] = repeat_values.get(odd, 0) + 1
        return repeat_values[odd]

    if is_valid_missing_odds_placeholder(name, attrs):
        return MISSING_PLACEHOLDER_VALUE
    
    if is_h2h_name(name):
        split_attr = split_attrs(attrs)
        index = 0
        field = ""

        for attr in split_attr:
            if attr.isdigit():
                index = int(attr)
            else:
                field = attr

        container = values.get(name, [])
        h2h_dict = container[index - 1].as_placeholder() if index <= len(container) else {field: "-"}
        return h2h_dict.get(field, "")

    return ""

def rich_parts(cell: Cell) -> list[tuple[str, InlineFont | None]]:
    if not isinstance(cell.value, CellRichText):
        return [(cell_text(cell.value), None)]

    parts: list[tuple[str, InlineFont | None]] = []
    for part in cell.value:
        if isinstance(part, TextBlock):
            parts.append((part.text, copy(part.font)))
        else:
            parts.append((str(part), None))
    return parts


def append_rich_text(result: CellRichText, font: InlineFont | None, text: Any) -> None:
    rendered_text = "" if text is None else str(text)
    if not rendered_text:
        return
    if font is None:
        result.append(rendered_text)
    else:
        result.append(TextBlock(copy(font), rendered_text))


def normalize_rich_text(result: CellRichText) -> CellRichText:
    normalized = CellRichText()
    pending_whitespace = ""

    for part in result:
        text = str(part)
        if text.isspace():
            pending_whitespace += text
            continue

        if pending_whitespace:
            if isinstance(part, TextBlock):
                part = TextBlock(copy(part.font), pending_whitespace + part.text)
            elif normalized and isinstance(normalized[-1], TextBlock):
                normalized[-1].text += pending_whitespace
            else:
                normalized.append(pending_whitespace)
            pending_whitespace = ""

        normalized.append(copy(part) if isinstance(part, TextBlock) else part)

    if pending_whitespace and normalized and isinstance(normalized[-1], TextBlock):
        normalized[-1].text += pending_whitespace

    normalized._opt()
    return normalized


def render_rich_cell(
    cell: Cell,
    values: dict[str, Any],
    worksheet: Worksheet,
) -> tuple[str | CellRichText | RenderedArrayFormula, Any | None]:
    if isinstance(cell.value, ArrayFormula):
        def replace(match: re.Match[str]) -> str:
            value = resolve_placeholder(values, match.group(1))
            return "" if value is None else str(value)

        return (
            RenderedArrayFormula(
                formula=PLACEHOLDER_RE.sub(replace, cell.value.text),
                array_ref=cell.value.ref,
            ),
            None,
        )

    result = CellRichText()
    has_rich_text = isinstance(cell.value, CellRichText)
    has_format_attr = False
    conditional_fill: Any | None = None

    for text, original_font in rich_parts(cell):
        position = 0
        for match in PLACEHOLDER_RE.finditer(text):
            append_rich_text(result, original_font, text[position:match.start()])

            placeholder, format_attr = split_format_attr(match.group(1))
            rendered_value = resolve_placeholder(values, placeholder)
            placeholder_font = original_font
            if format_attr is not None:
                format_cell = resolve_format_cell(format_attr, values, worksheet)
                placeholder_font = font_to_inline(format_cell.font)
                if conditional_fill is None:
                    conditional_fill = copy(format_cell.fill)
                has_format_attr = True
            append_rich_text(result, placeholder_font, rendered_value)
            position = match.end()

        append_rich_text(result, original_font, text[position:])

    if has_rich_text or has_format_attr:
        result._opt()
        return normalize_rich_text(result), conditional_fill
    rendered_value = str(result)
    return ("" if rendered_value.isspace() else rendered_value), conditional_fill

def render_rich_rows(
    source_rows: list[tuple[Cell, ...]],
    data_rows: list[dict[str, Any]],
    worksheet: Worksheet,
) -> list[list[tuple[str | CellRichText | RenderedArrayFormula, Any | None]]]:
    rendered: list[list[tuple[str | CellRichText | RenderedArrayFormula, Any | None]]] = []
    repeats = [{} for _ in range(max(len(row) for row in source_rows))]

    for values in data_rows:
        for source_row in source_rows:
            row: list[tuple[str | CellRichText | RenderedArrayFormula, Any | None]] = []
            for column, cell in enumerate(source_row):
                values["repeat"] = repeats[column]
                row.append(render_rich_cell(cell, values, worksheet))
            rendered.append(row)
    return rendered


def copy_cell_style(source: Cell, target: Cell) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = copy(source.number_format)
        target.protection = copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def template_rows(worksheet: Worksheet, start_row: int) -> list[tuple[Cell, ...]]:
    return [row for row in worksheet.iter_rows(min_row=start_row) if row_has_value(row)]


def clear_template_area(worksheet: Worksheet, start_row: int) -> None:
    if worksheet.max_row >= start_row:
        worksheet.delete_rows(start_row, worksheet.max_row - start_row + 1)


def read_xlsx_template(
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> XlsxTemplate:
    template_path = ensure_workbook_path(path)
    workbook = load_workbook(template_path, rich_text=True, keep_vba=keep_vba(template_path))
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active

        if header_row < 1:
            raise RuntimeError("Номер строки заголовков должен быть больше 0.")

        sheet_title = worksheet.title
        headers = [cell_text(cell.value) for cell in worksheet[header_row]]
        row_templates: list[list[Any]] = []

        for row in worksheet.iter_rows(min_row=header_row + 1):
            if not row_has_value(row):
                break
            row_templates.append([cell.value for cell in row])
    finally:
        close_workbook(workbook)

    if not any(headers):
        raise RuntimeError("В .xlsx шаблоне не найдена строка заголовков.")
    if not row_templates:
        raise RuntimeError("В .xlsx шаблоне не найдена строка с плейсхолдерами.")

    return XlsxTemplate(
        headers=headers,
        row_templates=row_templates,
        sheet_name=sheet_title,
        header_row=header_row,
        template_start_row=header_row + 1,
    )


def read_xlsx_sheet_templates(path: str | Path) -> list[tuple[XlsxTemplate, str]]:
    template_path = ensure_workbook_path(path)
    workbook = load_workbook(template_path, rich_text=True, keep_vba=keep_vba(template_path))
    try:
        sheet_configs: list[tuple[str, str]] = []
        for worksheet in workbook.worksheets:
            try:
                config_value = worksheet[worksheet.title].value
            except ValueError as error:
                raise RuntimeError(
                    f"Имя листа '{worksheet.title}' должно быть ссылкой на ячейку, например Z1."
                ) from error
            config_text = cell_text(config_value).strip()
            if not config_text:
                raise RuntimeError(
                    f"В ячейке {worksheet.title}!{worksheet.title} не найдены аргументы листа."
                )
            sheet_configs.append((worksheet.title, config_text))
    finally:
        close_workbook(workbook)

    return [
        (read_xlsx_template(template_path, sheet_name), config_text)
        for sheet_name, config_text in sheet_configs
    ]

def fill_xlsx_template(
    data_rows: list[dict[str, Any]],    
    template: XlsxTemplate,
    template_path: str | Path,
    output_path: str | Path,
    *,
    copy_template: bool = True,
    target_sheet_name: str | None = None,
) -> None:
    template_path = ensure_workbook_path(template_path)
    output_path = ensure_workbook_path(output_path)
    if template_path.suffix.lower() != output_path.suffix.lower():
        raise RuntimeError("Расширение выходного файла должно совпадать с расширением шаблона.")
    if template_path.resolve() == output_path.resolve():
        raise RuntimeError("Выходной файл не должен совпадать с файлом шаблона.")

    if copy_template:
        copy2(template_path, output_path)
    elif not output_path.exists():
        raise RuntimeError(f"Выходной файл не найден: {output_path}")
    workbook = load_workbook(output_path, rich_text=True, keep_vba=keep_vba(output_path))
    try:
        worksheet = workbook[template.sheet_name]
        start_row = template.header_row + 1
        source_rows = template_rows(worksheet, start_row)
        if not source_rows:
            raise RuntimeError("В .xlsx шаблоне не найдена строка с плейсхолдерами.")

        rendered_rows = render_rich_rows(source_rows, data_rows, worksheet)
        style_rows = [[cell for cell in row] for row in source_rows]

        array_formulas: dict[tuple[int, int], tuple[RenderedArrayFormula, str]] = {}
        array_cells: set[tuple[int, int]] = set()
        for row_offset, rendered_row in enumerate(rendered_rows):
            target_row_index = start_row + row_offset
            style_row = style_rows[row_offset % len(style_rows)]
            for column_index, (value, _) in enumerate(rendered_row, start=1):
                if not isinstance(value, RenderedArrayFormula):
                    continue
                row_offset_from_source = target_row_index - style_row[column_index - 1].row
                array_formulas[(target_row_index, column_index)] = (
                    value,
                    value.shifted_ref(row_offset_from_source),
                )
                array_cells.update(value.cells(row_offset_from_source))

        if keep_vba(output_path):
            source_row_numbers = [row[0].row for row in source_rows]
        else:
            clear_template_area(worksheet, start_row)

            for row_offset, rendered_row in enumerate(rendered_rows):
                target_row_index = start_row + row_offset
                style_row = style_rows[row_offset % len(style_rows)]
                for column_index, (value, conditional_fill) in enumerate(rendered_row, start=1):
                    target_cell = worksheet.cell(row=target_row_index, column=column_index)
                    if column_index <= len(style_row):
                        source_cell = style_row[column_index - 1]
                        copy_cell_style(source_cell, target_cell)
                    if (target_row_index, column_index) in array_cells:
                        continue
                    target_cell.value = value
                    if conditional_fill is not None:
                        target_cell.fill = conditional_fill

            for (row, column), (formula, array_ref) in array_formulas.items():
                worksheet.cell(row=row, column=column).value = ArrayFormula(ref=array_ref, text=formula.formula)

            if target_sheet_name is not None:
                worksheet.title = target_sheet_name
            workbook.save(output_path)
            return
    finally:
        close_workbook(workbook)

    fill_template_with_excel(
        output_path,
        template.sheet_name,
        start_row,
        source_row_numbers,
        rendered_rows,
        target_sheet_name,
    )
