from __future__ import annotations

import re
import sys
from contextlib import suppress
from pathlib import Path
from typing import Any

from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

try:
    import xlwings as xw
except ImportError:
    xw = None

XL_GENERAL_NUMBER_FORMAT_RU = "Общий"


def copied_formula_needs_no_rendering(source_cell: Any) -> bool:
    """Скопированная Excel-формула сама сдвигает ссылки при вставке строки."""
    if not source_cell.api.HasFormula:
        return False
    return "{" not in str(source_cell.api.Formula)


def shift_rendered_formula(
    formula: str,
    source_row: int,
    source_column: int,
    target_row: int,
    target_column: int,
) -> str:
    origin = f"{get_column_letter(source_column)}{source_row}"
    target = f"{get_column_letter(target_column)}{target_row}"
    return Translator(formula, origin=origin).translate_formula(target)


def write_excel_text(cell: Any, text: str) -> None:
    if text.startswith("="):
        application = cell.api.Application
        if cell.api.NumberFormat == "@":
            cell.api.NumberFormatLocal = XL_GENERAL_NUMBER_FORMAT_RU
        decimal_separator = application.DecimalSeparator
        formula = re.sub(r"(?<=\d)\.(?=\d)", decimal_separator, text)
        cell.api.FormulaLocal = formula
        return
    cell.api.NumberFormat = "@"
    cell.api.Value = text


def write_excel_text_range(cell_range: Any, values: list[str]) -> None:
    cell_range.api.NumberFormat = "@"
    if len(values) == 1:
        cell_range.api.Value = values[0]
    else:
        cell_range.api.Value = (tuple(values),)


def excel_rgb(color: Any) -> int | None:
    if color is None or color.type != "rgb" or not color.rgb:
        return None

    rgb = color.rgb[-6:]
    try:
        return int(rgb[4:6] + rgb[2:4] + rgb[:2], 16)
    except ValueError:
        return None


def apply_excel_inline_font(excel_font: Any, font: Any) -> None:
    for name, value in {
        "Name": font.rFont,
        "Size": font.sz,
        "Bold": font.b,
        "Italic": font.i,
    }.items():
        if value is not None:
            setattr(excel_font, name, value)

    if font.strike is not None:
        excel_font.Strikethrough = font.strike
    if font.vertAlign is not None:
        excel_font.Superscript = font.vertAlign == "superscript"
        excel_font.Subscript = font.vertAlign == "subscript"

    underline = {"single": 2, "double": -4119, "singleAccounting": 4, "doubleAccounting": 5}
    if font.u is not None:
        excel_font.Underline = underline.get(font.u, -4142)

    color = excel_rgb(font.color)
    if color is not None:
        excel_font.Color = color


def write_excel_rich_text(cell: Any, value: Any) -> None:
    text = str(value)
    write_excel_text(cell, text)
    if text.startswith("="):
        return

    position = 1
    for part in value:
        part_text = str(part)
        font = getattr(part, "font", None)
        if font is not None and part_text:
            excel_font = cell.api.GetCharacters(position, len(part_text)).Font
            apply_excel_inline_font(excel_font, font)
        position += len(part_text)


def apply_excel_fill(cell: Any, fill: Any) -> None:
    if fill.fill_type != "solid":
        return

    color = excel_rgb(fill.fgColor)
    if color is None:
        return

    cell.api.Interior.Pattern = 1
    cell.api.Interior.Color = color


def rename_sheets_with_excel(
    output_path: Path,
    sheet_names: dict[str, str],
) -> None:
    if sys.platform != "win32":
        raise RuntimeError("Переименование листов .xlsm выполняется только в Windows через Microsoft Excel.")
    if xw is None:
        raise RuntimeError("Для .xlsm установите xlwings: py -m pip install -r requirements.txt")

    changes = {
        source_name: target_name
        for source_name, target_name in sheet_names.items()
        if source_name != target_name
    }
    if not changes:
        return

    try:
        app = xw.App(visible=False, add_book=False)
    except Exception as error:
        raise RuntimeError(f"Не удалось запустить Microsoft Excel: {error}") from error

    app.display_alerts = False
    app.screen_updating = False
    workbook = None
    try:
        workbook = app.books.open(str(output_path), update_links=False, read_only=False)
        existing_names = {sheet.name for sheet in workbook.sheets}
        missing_names = set(changes) - existing_names
        if missing_names:
            raise RuntimeError(
                f"В книге не найдены листы: {', '.join(sorted(missing_names))}"
            )

        occupied_names = existing_names - set(changes)
        conflicts = set(changes.values()) & occupied_names
        if conflicts:
            raise RuntimeError(
                f"Имена уже заняты другими листами: {', '.join(sorted(conflicts))}"
            )

        temporary_names: dict[str, str] = {}
        reserved_names = existing_names | set(changes.values())
        temporary_index = 1
        for source_name in changes:
            temporary_name = f"__fsp_tmp_{temporary_index}__"
            while temporary_name in reserved_names:
                temporary_index += 1
                temporary_name = f"__fsp_tmp_{temporary_index}__"
            temporary_index += 1
            reserved_names.add(temporary_name)
            workbook.sheets[source_name].name = temporary_name
            temporary_names[temporary_name] = changes[source_name]

        for temporary_name, target_name in temporary_names.items():
            workbook.sheets[temporary_name].name = target_name

        workbook.save()
    except Exception as error:
        raise RuntimeError(f"Не удалось переименовать листы через Microsoft Excel: {error}") from error
    finally:
        try:
            if workbook is not None:
                workbook.close()
        finally:
            app.quit()


def fill_template_with_excel(
    output_path: Path,
    sheet_name: str,
    start_row: int,
    source_row_numbers: list[int],
    rendered_rows: list[list[tuple[Any, Any | None]]],
    target_sheet_name: str | None = None,
) -> None:
    if sys.platform != "win32":
        raise RuntimeError("Заполнение .xlsm с надстройками выполняется только в Windows через Microsoft Excel.")
    if xw is None:
        raise RuntimeError("Для .xlsm установите xlwings: py -m pip install -r requirements.txt")

    try:
        app = xw.App(visible=False, add_book=False)
    except Exception as error:
        raise RuntimeError(f"Не удалось запустить Microsoft Excel: {error}") from error

    app.display_alerts = False
    app.screen_updating = False
    previous_calculation = None
    previous_enable_events = None
    workbook = None
    source_sheet = None
    try:
        try:
            enable_events = app.api.EnableEvents
            app.api.EnableEvents = False
            previous_enable_events = enable_events
        except Exception as error:
            print(f"Excel: не удалось отключить события: {error}", flush=True)

        workbook = app.books.open(str(output_path), update_links=False, read_only=False)
        worksheet = workbook.sheets[sheet_name]

        try:
            calculation = app.calculation
            app.calculation = "manual"
            previous_calculation = calculation
        except Exception as error:
            print(f"Excel: не удалось отключить пересчёт: {error}", flush=True)

        worksheet.api.Copy(After=worksheet.api)
        source_sheet = workbook.sheets.active
        source_name = "__flashscore_template_source__"
        sheet_names = {sheet.name for sheet in workbook.sheets}
        suffix = 1
        while source_name in sheet_names:
            suffix += 1
            source_name = f"__flashscore_template_source_{suffix}__"
        source_sheet.name = source_name

        last_row = worksheet.used_range.last_cell.row
        if last_row >= start_row:
            worksheet.api.Rows(f"{start_row}:{last_row}").Delete()

        total_rows = len(rendered_rows)
        array_cells: set[tuple[int, int]] = set()
        for row_offset, rendered_row in enumerate(rendered_rows):
            target_row = start_row + row_offset
            source_row = source_row_numbers[row_offset % len(source_row_numbers)]
            for value, _ in rendered_row:
                if not hasattr(value, "array_ref"):
                    continue
                row_offset_from_source = target_row - source_row
                array_cells.update(value.cells(row_offset_from_source))

        source_formula_passthrough: dict[tuple[int, int], bool] = {}
        max_columns = max((len(row) for row in rendered_rows), default=0)
        for source_row in set(source_row_numbers):
            for column in range(1, max_columns + 1):
                source_cell = source_sheet.cells(source_row, column)
                source_formula_passthrough[(source_row, column)] = (
                    copied_formula_needs_no_rendering(source_cell)
                )

        print(f"Excel: заполнение {total_rows} строк...", flush=True)
        for row_offset, rendered_row in enumerate(rendered_rows):
            target_row = start_row + row_offset
            source_row = source_row_numbers[row_offset % len(source_row_numbers)]
            source_range = source_sheet.range((source_row, 1), (source_row, len(rendered_row)))
            target_range = worksheet.range((target_row, 1), (target_row, len(rendered_row)))
            source_range.api.Copy(Destination=target_range.api)

            plain_runs: list[tuple[int, list[str]]] = []
            plain_start = 0
            plain_values: list[str] = []
            individual_values: list[tuple[int, Any, bool]] = []
            fills: list[tuple[int, Any]] = []

            def finish_plain_run() -> None:
                nonlocal plain_start, plain_values
                if plain_values:
                    plain_runs.append((plain_start, plain_values))
                    plain_start = 0
                    plain_values = []

            for column, (value, conditional_fill) in enumerate(rendered_row, start=1):
                if (target_row, column) in array_cells:
                    finish_plain_run()
                    continue

                if conditional_fill is not None:
                    fills.append((column, conditional_fill))

                if source_formula_passthrough[(source_row, column)]:
                    finish_plain_run()
                    continue

                is_rich_text = hasattr(value, "_opt")
                rendered_value = "" if value is None else str(value)
                if rendered_value.startswith("="):
                    finish_plain_run()
                    rendered_value = shift_rendered_formula(
                        rendered_value,
                        source_row,
                        column,
                        target_row,
                        column,
                    )
                    individual_values.append((column, rendered_value, False))
                elif is_rich_text:
                    finish_plain_run()
                    individual_values.append((column, value, True))
                else:
                    if not plain_values:
                        plain_start = column
                    plain_values.append(rendered_value)

            finish_plain_run()

            for first_column, values in plain_runs:
                last_column = first_column + len(values) - 1
                text_range = worksheet.range(
                    (target_row, first_column),
                    (target_row, last_column),
                )
                try:
                    write_excel_text_range(text_range, values)
                except Exception as error:
                    raise RuntimeError(
                        f"Лист {worksheet.name}, диапазон {text_range.address}: {error}"
                    ) from error

            for column, value, is_rich_text in individual_values:
                target_cell = worksheet.cells(target_row, column)
                try:
                    if is_rich_text:
                        write_excel_rich_text(target_cell, value)
                    else:
                        write_excel_text(target_cell, value)
                except Exception as error:
                    raise RuntimeError(
                        f"Лист {worksheet.name}, ячейка {target_cell.address}: {error}"
                    ) from error

            for column, conditional_fill in fills:
                target_cell = worksheet.cells(target_row, column)
                try:
                    apply_excel_fill(target_cell, conditional_fill)
                except Exception as error:
                    raise RuntimeError(
                        f"Лист {worksheet.name}, ячейка {target_cell.address}: {error}"
                    ) from error

            if (row_offset + 1) % 100 == 0 or row_offset + 1 == total_rows:
                print(f"Excel: заполнено {row_offset + 1}/{total_rows} строк.", flush=True)

        source_sheet.delete()
        source_sheet = None
        if target_sheet_name is not None:
            worksheet.name = target_sheet_name
        print("Excel: сохранение файла...", flush=True)
        if previous_enable_events is not None:
            app.api.EnableEvents = previous_enable_events
            previous_enable_events = None
        if previous_calculation is not None:
            app.calculation = previous_calculation
            previous_calculation = None
        workbook.save()
        print("Excel: файл сохранён.", flush=True)
    except Exception as error:
        raise RuntimeError(f"Не удалось заполнить .xlsm через Microsoft Excel: {error}") from error
    finally:
        try:
            if source_sheet is not None:
                with suppress(Exception):
                    source_sheet.delete()
        finally:
            try:
                if workbook is not None:
                    workbook.close()
            finally:
                if previous_enable_events is not None:
                    with suppress(Exception):
                        app.api.EnableEvents = previous_enable_events
                if previous_calculation is not None:
                    with suppress(Exception):
                        app.calculation = previous_calculation
                app.quit()
